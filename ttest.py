

from difflib import SequenceMatcher, Differ
import openpyxl as xl
from openpyxl.styles import Color, PatternFill, Font
import xlsxwriter
# import pyexcel
import pandas
import xlrd
# from xls2xlsx import XLS2XLSX as xlxl

# https://xlsxwriter.readthedocs.io/example_rich_strings.html
# import difflib


class KoreanDiff:
    def __init__(self, source_path, compare_col):
        if source_path.split(".")[-1] == 'xls':
            # workbook = xl.Workbook(source_path)
            # source_path += 'x'
            # workbook.save(source_path)
            temp = source_path
            source_path += 'x'
            df = pandas.read_excel(temp)
            df.to_excel(source_path, index=False)
            # xls = pandas.ExcelFile(temp)

            # pyexcel.save_book_as(file_name=temp,
            #    dest_file_name=source_path)

        self.source_excel_path = source_path
        temp_target_path_list = source_path.split("\\")
        temp_target_path_list[-1] = "compared_" + temp_target_path_list[-1]
        self.target_excel_path = "\\".join(temp_target_path_list)

        

        self.source_row_char, self.target_row_char = compare_col

        self.source_row_num = ord(self.source_row_char) - 65
        self.target_row_num = ord(self.target_row_char) - 65

        self.compare_col = compare_col

        # self.red_color = Font(color='FE2E2E')
        # self.black_color = Font(color='000000')


        self.output_workbook = xlsxwriter.Workbook(self.target_excel_path)
        self.output_sheet = self.output_workbook.add_worksheet()

        self.cell_width = 40

        self.output_sheet.set_column(self.source_row_num, self.source_row_num, self.cell_width)
        self.output_sheet.set_column(self.target_row_num, self.target_row_num, self.cell_width)

        self.cell_format = self.output_workbook.add_format({'text_wrap': True})
        
        self.red_color = self.output_workbook.add_format({'color': 'red', 'text_wrap': True})
        self.black_color = self.output_workbook.add_format({'color': 'black', 'text_wrap': True})
        # write_ws.append([1,2,3])

        # print(f"source path: {self.source_excel_path}")
        # print(f"target path: {self.target_excel_path}")
        return

    def write_data_to_cell(self, pos, data, start, size):
        '''
        pos : tuple
            alphabet and number
        data : string
            input data
        start : list
            same block start index
        size : list
            same block size
        '''
        rendered_data = []
        loc = pos[0] + str(pos[1])
        # print(f'loc {loc}')
        # is_red_block = False
        # temp =  ['hello ',  self.red_color, 'this is bold_red',  'and default string']
        for i, point in enumerate(start):
            if i == 0:
                if point > 0: #start with diff 
                    rendered_data.append(self.red_color)
                    rendered_data.append(data[:point])
                    rendered_data.append(data[point:point + size[i]])
                else: # self.black_color
                    # rendered_data.append(self.black_color)
                    rendered_data.append(data[:size[i]])
            else:
                # if is_red_block:
                if point != point + size[i]:
                    rendered_data.append(data[point:point + size[i]])
                
            # next black
            # check end 
            if i < len(start) - 1:
                if point + size[i] < start[i + 1]:
                    rendered_data.append(self.red_color)
                if start[i + 1] != point + size[i]:
                    rendered_data.append(data[point + size[i]:start[i + 1]])
           
        print(rendered_data)
        if len(rendered_data) > 1:
            self.output_sheet.write_rich_string(loc, *rendered_data, self.cell_format)
        else:
            self.output_sheet.write(loc, rendered_data[0], self.cell_format)
        # self.output_sheet.write_rich_string(loc, 'abcd', self.black_color, 'abcd')
        return 

    def run(self):
        # read source excel
        self.wb = xl.load_workbook(self.source_excel_path)
        self.first_sheetname = self.wb.sheetnames[0]
        # do  get matching blocks
        for r_index, row in enumerate(self.wb[self.first_sheetname].iter_rows(min_row = 2)):
        # for each line
            # print(row[self.source_row_num].value)
            # print(row[self.target_row_num].value)
            
            # fill red all
            # row[self.source_row_num].fill = self.red_color
            # row[self.target_row_num].fill = self.red_color

            if not row[self.source_row_num].value:
                self.output_sheet.write(self.target_row_char + str(r_index + 1), row[self.target_row_num].value, self.red_color)
                continue
            if not row[self.target_row_num].value:
                self.output_sheet.write(self.source_row_char + str(r_index + 1), row[self.source_row_num].value, self.red_color)
                continue
            # same match -> black
            s = SequenceMatcher(None, row[self.source_row_num].value, row[self.target_row_num].value)
            # print(s.get_matching_blocks())
            source_index_list, target_index_list, size_list = [], [], []
            for mb in s.get_matching_blocks():
                # print(mb, row[self.source_row_num].value[mb.a:mb.a+mb.size])
                # print(mb, row[self.target_row_num].value[mb.b:mb.b+mb.size])
                source_index_list.append(mb.a)
                target_index_list.append(mb.b)
                size_list.append(mb.size)
            # print(f"row {r_index + 1}")
            self.write_data_to_cell((self.source_row_char, r_index + 1), row[self.source_row_num].value, source_index_list, size_list)
            self.write_data_to_cell((self.target_row_char, r_index + 1), row[self.target_row_num].value, target_index_list, size_list)
            print("*"*30)
            # break
        # self.write_data_to_cell(('A', 2), '', [], [])
        self.output_workbook.close()
        return 


if __name__ == "__main__":
    # dir_path = "./sample/"
    source_path = "D:\\codeSet\\pythonTest\\pyKoreanDiff\\sample\\Differential_Chart.xls"
    compare_col = ('B', 'G')
    
    # with open(source_path, 'rt', encoding='UTF8') as f:
    #     source = f.read()
    # with open(target_path, 'rt', encoding='UTF8') as f:
    #     target = f.read()
    # __run_diff(source, target)
    
    k_diff = KoreanDiff(source_path, compare_col)
    k_diff.run()
    pass