from difflib import SequenceMatcher
import openpyxl as xl
import xlsxwriter
import pandas
import tkinter.font 
from tkinter import filedialog as fd
import tkinter as tk

# https://xlsxwriter.readthedocs.io/example_rich_strings.html
# import difflib

class KoreanDiff:
    def __init__(self, source_path, compare_col, diffYes_col):
        # xls -> xlsx (xls 파일은 작업 불가능함)
        if source_path.split(".")[-1] == 'xls':
            temp = source_path
            source_path += 'x'
            df = pandas.read_excel(temp)
            df.to_excel(source_path, index=False)

        self.source_excel_path = source_path
        splitChar = '//' 
        # 파일 경로 일치시킴
        if '\\' not in source_path: 
            splitChar = '/'

        temp_target_path_list = source_path.split(splitChar)
        temp_target_path_list[-1] = "compared_" + temp_target_path_list[-1]
        self.target_excel_path = splitChar.join(temp_target_path_list)

        # compare column number
        self.source_row_char, self.target_row_char = compare_col
        self.diffYesCol = diffYes_col

        # compare column char
        self.source_row_num = ord(self.source_row_char) - 65
        self.target_row_num = ord(self.target_row_char) - 65
        self.diff_yes_num   = ord(self.diffYesCol) - 65

        # create workbook
        self.output_workbook = xlsxwriter.Workbook(self.target_excel_path)
        self.output_sheet = self.output_workbook.add_worksheet()

        self.cell_width = 40    # output cell width

        self.output_sheet.set_column(self.source_row_num, self.source_row_num, self.cell_width)
        self.output_sheet.set_column(self.target_row_num, self.target_row_num, self.cell_width)

        # cell background color setting
        self.change_bg_color = '#FFFFCC'
        self.head_bg_color = '#FF6600'

        self.rem_word = '_x000D_'

        # text auto newline
        self.cell_format = self.output_workbook.add_format({'text_wrap': True, 'bg_color': self.change_bg_color, 'border': 1})
        self.cell_head = self.output_workbook.add_format({'text_wrap': True, 'bg_color': self.head_bg_color, 'border': 1})
        # text color
        self.red_color = self.output_workbook.add_format({'color': 'red', 'text_wrap': True, 'bg_color': self.change_bg_color, 'border': 1})
        self.black_color = self.output_workbook.add_format({'color': 'black', 'text_wrap': True, 'bg_color': self.change_bg_color, 'border': 1})
        self.orange_color = self.output_workbook.add_format({'color': 'orange', 'text_wrap': True, 'bg_color': self.change_bg_color, 'border': 1})
        self.green_color = self.output_workbook.add_format({'color': 'green', 'text_wrap': True, 'bg_color': self.change_bg_color, 'border': 1})
        return

    def write_data_to_cell(self, pos, data, start, size):
        '''
        pos : tuple
            alphabet and number
        data : string
            input data
        start : list
            same block start index
        size : list
            same block size
        '''
        rendered_data = []
        start_with = 'black'
        loc = pos[0] + str(pos[1])  # cell location

        for i, point in enumerate(start):
            if i == 0:  # if first data
                if point > 0: #start with red
                    start_with = 'red'
                    rendered_data.append(self.red_color)
                    rendered_data.append(data[:point])
                    if size[i] > 0: # if black data size > 0
                        rendered_data.append(self.black_color)
                        rendered_data.append(data[point:point + size[i]])
                else: # start with black
                    rendered_data.append(self.black_color)
                    rendered_data.append(data[:size[i]])
                    
            else:
                if size[i] > 0: # input black data
                    rendered_data.append(data[point:point + size[i]])
                
            # next black
            # 다음 red block 연결
            if i < len(start) - 1:  # if not last data
                # if red exist
                if point + size[i] < start[i + 1]:  
                    rendered_data.append(self.red_color)
                    rendered_data.append(data[point + size[i]:start[i + 1]])
           
        if len(rendered_data) > 2:  # over two block
            self.output_sheet.write_rich_string(loc, *rendered_data, self.cell_format)
        else:   # just single block
            if start_with == 'red':
                self.output_sheet.write(loc, rendered_data[1], self.red_color)
            else:
                self.output_sheet.write(loc, rendered_data[1], self.black_color)

        return 

    def run(self):
        # read source excel
        self.wb = xl.load_workbook(self.source_excel_path)
        self.first_sheetname = self.wb.sheetnames[0]

        # do  get matching blocks
        for r_index, row in enumerate(self.wb[self.first_sheetname].iter_rows(min_row = 0)):
            # 피쳐명인 경우
            if not row[0].value and row[self.source_row_num].value:
                self.output_sheet.write(self.source_row_char + str(r_index + 1), row[self.source_row_num].value, self.cell_head)
                continue

            # if no data
            if not row[self.source_row_num].value:  
                # another data to red all
                if row[self.target_row_num].value: 
                    row[self.target_row_num].value = row[self.target_row_num].value.replace(self.rem_word, '')
                    self.output_sheet.write(self.target_row_char + str(r_index + 1), row[self.target_row_num].value, self.red_color)
                continue
            if not row[self.target_row_num].value:
                if row[self.source_row_num].value: 
                    row[self.source_row_num].value = row[self.source_row_num].value.replace(self.rem_word, '')
                    self.output_sheet.write(self.source_row_char + str(r_index + 1), row[self.source_row_num].value, self.red_color)
                continue
            

            # remove trash word
            row[self.source_row_num].value = row[self.source_row_num].value.replace(self.rem_word, '')
            row[self.target_row_num].value = row[self.target_row_num].value.replace(self.rem_word, '')
            
            # 둘이 같은데 y 인 경우 -> continue
            # 둘이 같은데 모델 확인해야하는 경우 -> continue
            if row[self.source_row_num].value == row[self.target_row_num].value:
                font_color = self.black_color
                if '[모델사양' in row[self.target_row_num].value: # 모델사양 확인
                    font_color = self.orange_color
                elif row[self.diff_yes_num].value == "Y": # 첨부파일의 변경
                    font_color = self.green_color
                else: # 변경점 없음
                    pass

                self.output_sheet.write(self.source_row_char + str(r_index + 1), row[self.source_row_num].value, font_color)
                self.output_sheet.write(self.target_row_char + str(r_index + 1), row[self.target_row_num].value, font_color)
                continue
            
            s = SequenceMatcher(None, row[self.source_row_num].value, row[self.target_row_num].value)
            source_index_list, target_index_list, size_list = [], [], []
            for mb in s.get_matching_blocks():
                source_index_list.append(mb.a)
                target_index_list.append(mb.b)
                size_list.append(mb.size)
            
            self.write_data_to_cell((self.source_row_char, r_index + 1), row[self.source_row_num].value, source_index_list, size_list)
            self.write_data_to_cell((self.target_row_char, r_index + 1), row[self.target_row_num].value, target_index_list, size_list)
            
        self.output_workbook.close()

        return True

class FeatureCompare:
    def __init__(self):
        self.source_dir_path = "[Difference Chart Directory Path]"
        self.errorMsg = ""
        self.processPercent = 0
        self.compare_column = ['B', 'G', 'K']

        title = "Feature Compare v1.0"
        subtitle = 'PRMS Feature Compare'
        comment = """
1. Open -> PRMS에서 Export한 Feature Compare File 선택
2. Feature Compare에서 비교하고 싶은 두 Column을 입력 (기본값은 B, G)
3. 세번째 칸에 Y (변경됨)가 적힌 Column을 입력 (기본값은 K)
4. Compare 버튼을 통해 실행
5. 작업 문서와 동일한 폴더에 'compared_' 로 시작하는 결과 파일이 생성됩니다.

(※ 작업 폴더에 동일한 이름의 xlsx 파일이 생성될 수 있습니다)
        """
        copyrighter = "문의 : @minku.koo"
        
        self.window = tk.Tk()
        self.window.title(title)
        self.window.geometry('600x600')

        subtitleFont = tkinter.font.Font(size = 18, weight = 'bold')
        entryFont = tkinter.font.Font(size = 13)
        copyrightFont = tkinter.font.Font(size = 8)

        
        subtitleBox = tk.Label(self.window,
                        text=subtitle,
                        width = 600,
                        pady = 10,
                        font=subtitleFont,
                        justify='center')
        subtitleBox.pack()

        label = tk.Label(self.window,
                        text=comment,
                        width = 600,
                        padx = 20,
                        justify='left')
        label.pack()
        
        btnBorderFrame = tkinter.LabelFrame(self.window, bd = 2, bg = "black")
        btnBorderFrame.pack()

        fileOpenBtn = tkinter.Button(btnBorderFrame,
                                    font=entryFont,
                                    text = 'Open',
                                    padx = 10,
                                    pady=10,
                                    width = 16,
                                    command= self.__fileOpen
                                    )
        fileOpenBtn.pack()

        self.sourceDirPath = tk.Label(self.window,
                        text=self.source_dir_path,
                        width = 600,
                        padx = 20,
                        pady=20)
        self.sourceDirPath.pack()

        self.compareColEntryA = tkinter.Entry(
                                        self.window,
                                        width=6,
                                        font=entryFont
                                    )
        self.compareColEntryA.insert(0, self.compare_column[0])
        self.compareColEntryA.pack()

        self.compareColEntryB = tkinter.Entry(
                                        self.window,
                                        width=6,
                                        font=entryFont
                                    )
        self.compareColEntryB.insert(0, self.compare_column[1])
        self.compareColEntryB.pack(pady=6)

        self.diffYesCol = tkinter.Entry(
                                        self.window,
                                        width=6,
                                        font=entryFont
                                    )
        self.diffYesCol.insert(0, self.compare_column[2])
        self.diffYesCol.pack(pady=12)
        
        btnBorderFrame = tkinter.LabelFrame(self.window, bd = 2, bg = "black")
        btnBorderFrame.pack()

        compareBtn = tkinter.Button(btnBorderFrame,
                                    font=entryFont,
                                    text = 'Compare',
                                    padx = 10,
                                    pady=10,
                                    width = 16,
                                    command = self.__compareText
                                    )
        compareBtn.pack()

        self.progressBar = tk.Label(self.window,
                        text="Please, open the file",
                        width = 600,
                        padx = 20,
                        pady=20,
                        fg='black',
                        font=entryFont)
        self.progressBar.pack()

        copyrightLab = tk.Label(self.window,
                        fg='#0033CC',
                        text=copyrighter,
                        font=copyrightFont)
        copyrightLab.pack()
        
        self.window.mainloop()
        return 

    def __fileOpen(self):
        '''
        작업 파일 Open
        '''
        filepath = fd.askopenfilename(filetypes=[("Excel file","*.xlsx"),("Excel file", "*.xls")])

        if filepath:
            self.source_dir_path = filepath
            self.progressBar.config(text = "Press [Compare]", fg = 'black')
            self.sourceDirPath.config(text = self.source_dir_path)

        return 

    def __compareText(self):
        '''
        - Compare 버튼을 누르면 동작하는 event 함수
        '''
        k_diff = KoreanDiff(self.source_dir_path, (self.compareColEntryA.get(), self.compareColEntryB.get()), self.diffYesCol.get())
        check = False

        try:
            check = k_diff.run() # 컴페어 실행
        except IndexError as e:
            # 컴페어 실패할 경우 (컬럼에 데이터가 없는 경우)
            print(e)
            self.progressBar.config(text = "Error: Column data is not in Excel file !!", fg = 'red')
        
        except:
            self.progressBar.config(text = "Error !!", fg = 'red')

        if check:
            # 실행 성공
            self.progressBar.config(text = "Feature Compare Done", fg = 'blue')

        return 


if __name__ == "__main__":
    fc = FeatureCompare()